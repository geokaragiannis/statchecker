file="Ch03 - Oil.xlsx"
N=500  ##limit rows for testing



import openpyxl
import numpy as np
import pandas as pd
import os,os.path
#import pixiedust
import win32com.client
import re
import json

header=['Look-up value',
 'Look-up year',
 'Scenario',
 'Chapter heading',
 'Text',
 'Published value',
 'Author calculation',
 'Unit',
 'Source or Contact',
 'Comment',
 None,
 'Checker #1',
 'zakia',
 None,
 'Checker #2 ',
 'Claudia',
 None,
 'Checker #3 ']
 
 
 
col2ind=dict(zip(header,range(len(header))))



"""Iterate thorugh each cell and store values in a list."""



book=openpyxl.load_workbook(file)
book_val=openpyxl.load_workbook(file,data_only=True)

i=0
df=[]

for sheetname in book.sheetnames:
    
    sheet=book[sheetname]          # sheet to get formula
    sheet_val=book_val[sheetname]  #sheet to get value of formula
    
    text=PublishedValue=AuthorCalc_eq=AuthorCalc_Value=LookUpValue=LookUpYear=Scenario=None

    
    for row in sheet.rows:
    
        cells=[cell for cell in row]
        values=[cell.value for cell in row]
        
        
        if(not check_None(values)):
            
            if(values[col2ind['Text']]):            #check first with text, new text means other new values
                text=values[col2ind['Text']]
                PublishedValue=AuthorCalc_eq=AuthorCalc_Value=LookUpValue=LookUpYear=Scenario=None
            
            if(values[col2ind['Published value']]):   #new published value means new calculations
                PublishedValue=values[col2ind['Published value']]
                AuthorCalc_eq=AuthorCalc_Value=None
        
            if(values[col2ind['Author calculation']]):# and not isinstance(values[col2ind['Author calculation']],bool)):
                AuthorCalc_eq=values[col2ind['Author calculation']]
                AuthorCalc_Value=sheet_val[cells[col2ind['Author calculation']].coordinate].value
            
            
           # if(isinstance(AuthorCalc_eq,str) and 'LOOKUP' in AuthorCalc_eq ):
           #     lookup=1
                
                
            if(values[col2ind['Look-up value']]):
                LookUpValue=values[col2ind['Look-up value']]
                
                
            if(values[col2ind['Look-up year']]):
                LookUpYear=values[col2ind['Look-up year']]
                
                
            if(values[col2ind['Scenario']]):
                Scenario=values[col2ind['Scenario']]
                
    
                
        if(text and PublishedValue and AuthorCalc_eq and AuthorCalc_Value):
            #print([text,PublishedValue,AuthorCalc_eq,AuthorCalc_Value,LookUpValue,LookUpYear,Scenario])
            df.append([text,PublishedValue,AuthorCalc_eq,AuthorCalc_Value,LookUpValue,LookUpYear,Scenario])
            AuthorCalc_eq=AuthorCalc_Value=None
            i+=1
            print(i)
            if(i==N):
                break

    if(i==N):
        break
df[0][2]='Author Calculation Equations'
df[0][3]='Author calc Value'
            
claims=pd.DataFrame(df[1:],columns=df[0]) #remove first row and make it header


