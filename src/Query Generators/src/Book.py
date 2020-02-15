import openpyxl
from src.Sheet import Sheet

class Book:

    """

    Instance Variables
    ==================

    -book: openpyxl workbook object
    -book_val: openpyxl workbook object with values
    -file: filename
    -sheetnames: names of sheets of the workbook
    -sheet_dict: dict where sheetname is the key and the value is the Sheet object,intialized with None


    Methods
    =======

    -print_tabs: prints the sheetnames of the workbook



    -get_sheet: 

     Args: sheetname
     Returns: Sheet object

    """


    
    cached_books={}    
    def __init__(self,file):

        if(file in Book.cached_books):
            print("Loading file from cache.")
            self.book=Book.cached_books[file][0]
            self.book_val=Book.cached_books[file][1]

        else:
            print("Loading {}.".format(file))
            self.book=openpyxl.load_workbook(file)
            self.book_val=openpyxl.load_workbook(file,data_only=True)
            Book.cached_books[file]=[self.book,self.book_val]
            print("Finished loading {}.".format(file))
        self.file=file
        self.sheetnames=(self.book).sheetnames
        self.sheets_dict={sheetname:None for sheetname in self.sheetnames}
        
    
    def print_tabs(self):
        print(self.sheetnames)
    
    def get_sheet(self,sheetname):
        if(sheetname not in self.sheetnames):
            print("{} not found in sheets.".format(sheetnames))
            return
        if(not self.sheets_dict[sheetname]):
            self.sheets_dict[sheetname]=Sheet(self,sheetname)
        return self.sheets_dict[sheetname]
        
    
